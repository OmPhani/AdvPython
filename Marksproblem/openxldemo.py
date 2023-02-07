import openpyxl
wb = openpyxl.load_workbook("C://Users/user790/Desktop/marks_problem.xlsx")
sheet = wb.active
data = sheet['A3'].value
data1 = sheet['A1:A5']
data2 = sheet.cell(row=2, column=3).value
print(data)
print(data1)
print(data2)

for row in sheet.rows:
    print([data.value for data in row])
for i in range(1,12):
    print(sheet.cell(row =i,column = 1).value)